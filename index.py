import requests;
import json;
from openpyxl import Workbook;
from openpyxl.utils import get_column_letter;
import time;
from datetime import datetime;
from datetime import date;
from openpyxl import load_workbook;
import itertools, collections
import operator

infoFile = {}
with open("info.txt") as myfile:
    for line in myfile:
        name, var = line.partition(":")[::2]
        infoFile[name.strip()] = var

# print(infoFile['API'])

api = infoFile['API']

# quit()

# print("Put your input file in the same folder as this program...");
# api = input("Please enter API Key: ")

addressBook = load_workbook(filename = "input.xlsx");
addressCount = len(addressBook['Addresses']['A']);
currentTime = time.time();

buy24Hours = [];
buy48Hours = [];
buy72Hours = [];

sell24Hours = [];
sell48Hours = [];
sell72Hours = [];


b24h = []
b48h = []
b72h = []
s24h = []
s48h = []
s72h = []

rAll = [];

Tokens = {};

addresn = 1;


def verifyAddress(address):
    verfiy_api = f'https://api.bscscan.com/api?module=contract&action=getsourcecode&address={address}&apikey={api}';
    verify = requests.get(verfiy_api);
    verify = json.loads(verify.content);
    verify = verify["result"]
    # print(type(verify))
    # print(verify)
    # verify[len(verify) - 1] = '';
    # print(verify[0])
    # print(verfiy_api)
    verified = 0;
    for key in verify:
        # print(key.get('ContractName'))
   
        if(key.get('ContractName') == "PancakePair"):
            verified = 1;
        else:
            continue;
    # print (verfiy_api)
    return verified;

def calculateOverview(address, t, n, res):
    
    # if(Tokens[t]['Added'] == True and Tokens[t]['ACount'] == addresn):
    #     print(Tokens[t]['Added'])
    #     continue;
    # else:
    #     Tokens[t]['ACount'] == addresn;
    
    # t = t['tokenName']
    checkT = Tokens.get(t)
    if(checkT is None):
        Tokens[t] = {
            'Token': t,
            'Hour24': 0,
            'Hour48': 0,
            'Hour72': 0,
            'sHour24': 0,
            'sHour48': 0,
            'sHour72': 0,
            'Added': False,
            'ACount': n
        }
    else:
        return;
        # if(Tokens[t]['ACount'] != n):
        #     # print("c: ", n)
        #     Tokens[t]['ACount'] += 1;
        # else:
        #     return;

    

    # if(Tokens[t]['ACount'] == n and Tokens[t]['Added'] == True):
    #     return;

    # print(buy24Hours)

    # b24h = list(itertools.chain(*buy24Hours))
    # b48h = list(itertools.chain(*buy48Hours))
    # b72h = list(itertools.chain(*buy72Hours))
    
    # s24h = list(itertools.chain(*sell24Hours))
    # s48h = list(itertools.chain(*sell48Hours))
    # s72h = list(itertools.chain(*sell72Hours))

    

    # for te in res:

    #     timeS = ( currentTime - int(te['timeStamp']) ) / 3600;

    #     if(timeS > 72):
    #         # print("skip because of time")
    #         continue;
    #     else:
    #         if te['verifyAddressFrom'] == 0:
    #             if te['verifyAddressTo'] == 0:
    #                 # print("skip because of address")
    #                 continue;
        
    Tokens[t]['Hour24'] = b24h.count(t)
    Tokens[t]['Hour48'] = b48h.count(t)
    Tokens[t]['Hour72'] = b72h.count(t)

    Tokens[t]['sHour24'] = s24h.count(t)
    Tokens[t]['sHour48'] = s48h.count(t)
    Tokens[t]['sHour72'] = s72h.count(t)


        # if(te['to'] == address and te['tokenName'] == t ):
        #     if(timeS <= 24):
        #         Tokens[t]['Hour24'] += 1
        #     elif (timeS <= 48):
        #         Tokens[t]['Hour48'] += 1
        #     elif(timeS <= 72):
        #         Tokens[t]['Hour72'] += 1
        # if (te['from'] == address and te['tokenName'] == t ):
        #     if(timeS <= 24):
        #         Tokens[t]['sHour24'] += 1
        #     elif (timeS <= 48):
        #         Tokens[t]['sHour48'] += 1
        #     elif(timeS <= 72):
        #         Tokens[t]['sHour72'] += 1
    Tokens[t]['Added'] = True;
    # print(Tokens[t])


def calculateTransactions(res, address):
    timeS = ( currentTime - int(res['timeStamp']) ) / 3600;

    if(timeS > 72):
        # print("skip because of time")
        return;
    else:
        if res['verifyAddressFrom'] == 0:
            if res['verifyAddressTo'] == 0:
                # print("skip because of address")
                return;

    if(res['to'] == address):
        if(timeS <= 24):
            b24h.append(res['tokenName'])
            buy24Hours.append({
                'Address': res['to'],
                'Token': res['tokenName'],
                'Ammount': res['value'],
            })
        elif (timeS <= 48):
            b48h.append(res['tokenName'])
            buy48Hours.append({
                'Address': res['to'],
                'Token': res['tokenName'],
                'Ammount': res['value']
            })
        elif(timeS <= 72):
            b72h.append(res['tokenName'])
            buy72Hours.append({
                'Address': res['to'],
                'Token': res['tokenName'],
                'Ammount': res['value']
            })
    elif(res['from'] == address):
        if(timeS <= 24):
            s24h.append(res['tokenName'])
            sell24Hours.append({
                'Address': res['from'],
                'Token': res['tokenName'],
                'Ammount': res['value']
            })
        elif (timeS <= 48):
            s48h.append(res['tokenName'])
            sell48Hours.append({
                'Address': res['from'],
                'Token': res['tokenName'],
                'Ammount': res['value']
            })
        elif(timeS <= 72):
            s72h.append(res['tokenName'])
            sell72Hours.append({
                'Address': res['from'],
                'Token': res['tokenName'],
                'Ammount': res['value']
            })


# Sheet Settings
sheet = Workbook();
w0 = sheet.active;

w0.title = "Buy Overview";
w0.cell(column=1, row=1, value="Token");
w0.cell(column=2, row=1, value="24 Hours");
w0.cell(column=3, row=1, value="48 Hours");
w0.cell(column=4, row=1, value="72 Hours");

w1 = sheet.create_sheet(title="Sell Overview");
w1.cell(column=1, row=1, value="Token");
w1.cell(column=2, row=1, value="24 Hours");
w1.cell(column=3, row=1, value="48 Hours");
w1.cell(column=4, row=1, value="72 Hours");

w2 = sheet.create_sheet(title="24 Hour Buys");
w2.cell(column=1, row=1, value="Address");
w2.cell(column=2, row=1, value="Token");
w2.cell(column=3, row=1, value="Ammount");

w3 = sheet.create_sheet(title="24 Hour Sells");
w3.cell(column=1, row=1, value="Address");
w3.cell(column=2, row=1, value="Token");
w3.cell(column=3, row=1, value="Ammount");

w4 = sheet.create_sheet(title="48 Hour Buys");
w4.cell(column=1, row=1, value="Address");
w4.cell(column=2, row=1, value="Token");
w4.cell(column=3, row=1, value="Ammount");

w5 = sheet.create_sheet(title="48 Hour Sells");
w5.cell(column=1, row=1, value="Address");
w5.cell(column=2, row=1, value="Token");
w5.cell(column=3, row=1, value="Ammount");

w6 = sheet.create_sheet(title="72 Hour Buys");
w6.cell(column=1, row=1, value="Address");
w6.cell(column=2, row=1, value="Token");
w6.cell(column=3, row=1, value="Ammount");

w7 = sheet.create_sheet(title="72 Hour Sells");
w7.cell(column=1, row=1, value="Address");
w7.cell(column=2, row=1, value="Token");
w7.cell(column=3, row=1, value="Ammount");


for address in addressBook['Addresses']['A']:
    
    # print("Address Count: ", addresn)

    address = address.value.lower();

    print("Gettting data for address: ", address);

    # API Connection Information
    api_url = f'https://api.bscscan.com/api?module=account&action=tokentx&address={address}&sort=desc&apikey={api}';
    
    # API Request
    request = requests.get(api_url);

    # API Response
    data = json.loads(request.content);
    result = data["result"];
    # print( api_url );

    # Data Arrays
    buy = [];
    sell = [];

    checked = [];

    # Data Manipulation
    tno = 0;
    for res in result:
        timeStamp = ( currentTime - int(res['timeStamp']) ) / 3600;
        tno += 1;
        if(timeStamp > 72):
            # print("skipping")
            res.update({'verifyAddressFrom': 0});
            res.update({'verifyAddressTo': 0});
            print("Verifying Pancake Swap, Transaction #: ", tno, " - Skipping")
            continue;
        else:
            verifyAddressFrom = verifyAddress(res['from'])
            verifyAddressTo = verifyAddress(res['to'])
            res.update({'verifyAddressFrom': verifyAddressFrom});
            res.update({'verifyAddressTo': verifyAddressTo});
            calculateTransactions(res, address)
            print("Verifying Pancake Swap, Transaction #: ", tno)
            rAll.append(res);

        # print("v: ", verifyAddressFrom)
        

    
    addresn += 1;
for res in rAll :
    calculateOverview(address, res['tokenName'], addresn, rAll)

        # for r in reversed(result):
        #     timeStamp = ( currentTime - int(r['timeStamp']) ) / 3600;
        #     if(r['tokenName'] == t):
        #         if(r['to'] == address):
        #             if(timeStamp <= 24):                        
        #                 # tokens[t]['Hour24'] += 1
        #                 print(t," 24 count: ", tokens[t]['Hour24'])
        #             elif (timeStamp <= 48):
        #                 # tokens[t]['Hour48'] += 1
        #                 print(t," 48 count: ", tokens[t]['Hour24'])
        #             elif(timeStamp <= 72):
        #                 # tokens[t]['Hour72'] += 1

        #                 print(t," 24 count: ", tokens[t]['Hour24'])
        #             else:
        #                 continue;
        #         elif(r['from'] == address):

        #             if(timeStamp <= 24):
        #                 tokens[t]['sHour24'] += 1
        #             elif (timeStamp <= 48):
        #                 tokens[t]['sHour48'] += 1
        #             elif(timeStamp <= 72):
        #                 tokens[t]['sHour72'] += 1
        #             else:
        #                 continue;
    
i=2;
for r in buy24Hours:
    w2.cell(column=1, row=i, value=r['Address']);
    w2.cell(column=2, row=i, value=r['Token']);
    w2.cell(column=3, row=i, value=r['Ammount']);
    i += 1;
i=2;
for r in sell24Hours:
    w3.cell(column=1, row=i, value=r['Address']);
    w3.cell(column=2, row=i, value=r['Token']);
    w3.cell(column=3, row=i, value=r['Ammount']);
    i += 1;
i=2;
for r in buy48Hours:
    w4.cell(column=1, row=i, value=r['Address']);
    w4.cell(column=2, row=i, value=r['Token']);
    w4.cell(column=3, row=i, value=r['Ammount']);
    i += 1;
i=2;
for r in sell48Hours:
    w5.cell(column=1, row=i, value=r['Address']);
    w5.cell(column=2, row=i, value=r['Token']);
    w5.cell(column=3, row=i, value=r['Ammount']);
    i += 1;
i=2;
for r in buy72Hours:
    w6.cell(column=1, row=i, value=r['Address']);
    w6.cell(column=2, row=i, value=r['Token']);
    w6.cell(column=3, row=i, value=r['Ammount']);
    i += 1;
i=2;
for r in sell72Hours:
    w7.cell(column=1, row=i, value=r['Address']);
    w7.cell(column=2, row=i, value=r['Token']);
    w7.cell(column=3, row=i, value=r['Ammount']);
    i += 1;
# print(Tokens)


j=2;
for r in Tokens:
    if(Tokens[r]['Hour24'] == 0 and Tokens[r]['Hour48'] == 0 and Tokens[r]['Hour72'] == 0):
        continue;
    else:
        w0.cell(column=1, row=j, value=Tokens[r]['Token']);
        w0.cell(column=2, row=j, value=Tokens[r]['Hour24']);
        w0.cell(column=3, row=j, value=Tokens[r]['Hour48']);
        w0.cell(column=4, row=j, value=Tokens[r]['Hour72']);
        j += 1;

i=2;
for r in Tokens:    
    if(Tokens[r]['sHour24'] == 0 and Tokens[r]['sHour48'] == 0 and Tokens[r]['sHour72'] == 0):
        continue;
    else:
        w1.cell(column=1, row=i, value=Tokens[r]['Token']);
        w1.cell(column=2, row=i, value=Tokens[r]['sHour24']);
        w1.cell(column=3, row=i, value=Tokens[r]['sHour48']);
        w1.cell(column=4, row=i, value=Tokens[r]['sHour72']);
        i += 1;
w0.column_dimensions['A'].width = 25;
w1.column_dimensions['A'].width = 25;
# print(buy72Hours)
# print(buy72Hours)
print("Crawling Completed!")
sheet.save(filename = "output.xlsx");

